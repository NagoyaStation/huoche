#!/usr/bin/env python3
"""生成完整的 Excel 配置表，覆盖雪国列车全部游戏数值"""
import os, sys

# 确保 openpyxl 可用
try:
    import openpyxl
except ImportError:
    os.system(sys.executable + " -m pip install openpyxl --break-system-packages -q")
    import openpyxl

from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

OUT_DIR = os.path.join(os.path.dirname(os.path.abspath(__file__)), "..", "docs", "configs")
os.makedirs(OUT_DIR, exist_ok=True)

# ---------- 样式 ----------
HEADER_FONT = Font(bold=True, size=11)
HEADER_FILL = PatternFill("solid", fgColor="D9E1F2")
THIN_BORDER = Border(
    left=Side(style="thin"), right=Side(style="thin"),
    top=Side(style="thin"), bottom=Side(style="thin"),
)

def style_header(ws, row=1, ncols=4):
    for col in range(1, ncols + 1):
        cell = ws.cell(row=row, column=col)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.border = THIN_BORDER
        cell.alignment = Alignment(horizontal="center")

def auto_width(ws):
    for col in ws.columns:
        mx = 0
        letter = col[0].column_letter
        for cell in col:
            if cell.value is not None:
                mx = max(mx, len(str(cell.value)))
        ws.column_dimensions[letter].width = min(60, max(12, mx + 4))

# =====================================================================
# 1. GameConfig.xlsx — 参数表(#Sheet)
# =====================================================================
def make_game_config():
    wb = Workbook()
    # --- #Scroll ---
    ws = wb.active
    ws.title = "#Scroll"
    rows = [
        ("int",   "BASE_SCROLL_SPEED",       30,  "基础滚动速度 px/s"),
        ("float", "SCROLL_SPEED_PER_LEVEL",  1.5, "每级加速"),
        ("float", "PATH_WIDTH_RATIO",        0.30,"路径宽度占屏幕比例"),
    ]
    for r in rows: ws.append(r)
    auto_width(ws)

    # --- #Train ---
    ws2 = wb.create_sheet("#Train")
    rows2 = [
        ("int", "MAX_HP",     100, "列车最大生命值"),
        ("int", "HP_REGEN",   0,   "列车每秒自动回复"),
    ]
    for r in rows2: ws2.append(r)
    auto_width(ws2)

    # --- #Player ---
    ws3 = wb.create_sheet("#Player")
    rows3 = [
        ("int",   "SPEED",             185,  "玩家移动速度"),
        ("int",   "W",                 22,   "玩家宽度"),
        ("int",   "H",                 32,   "玩家高度"),
        ("int",   "MAX_CARRY",         10,   "最大携带数"),
        ("int",   "AUTO_ATTACK_RANGE", 48,   "自动攻击/采集范围"),
        ("float", "AUTO_ATTACK_INTERVAL", 0.35, "攻击间隔(秒)"),
        ("int",   "ATK",               30,   "玩家基础攻击力"),
        ("int",   "ATK_ZOMBIE",        30,   "对丧尸攻击力"),
    ]
    for r in rows3: ws3.append(r)
    auto_width(ws3)

    # --- #Zombie ---
    ws4 = wb.create_sheet("#Zombie")
    rows4 = [
        ("float", "SPAWN_INTERVAL",       4.0,  "基础丧尸生成间隔(秒)"),
        ("float", "SPAWN_INTERVAL_MIN",   1.0,  "最小生成间隔(秒)"),
        ("float", "SPAWN_INTERVAL_REDUCE",0.3,  "每级减少间隔(秒)"),
        ("int",   "SPEED",                30,   "丧尸移动速度 px/s"),
        ("float", "SPEED_PER_LEVEL",      1.5,  "每级加速"),
        ("int",   "SIZE",                 20,   "碰撞半径"),
        ("int",   "HP",                   3,    "丧尸基础生命值(旧,实际用zombieType)"),
        ("int",   "DAMAGE",               8,    "对列车伤害"),
        ("float", "ATK_INTERVAL",         1.2,  "攻击列车间隔"),
        ("int",   "MAX_BASE",             8,    "基础最大丧尸数"),
        ("int",   "MAX_PER_LEVEL",        2,    "每级增加最大丧尸数"),
        ("int",   "MAX_CAP",              30,   "绝对上限"),
        ("int",   "TYPE1_BASE_HP",        80,   "白T恤僵尸基础HP"),
        ("int",   "TYPE2_BASE_HP",        100,  "棕外套僵尸基础HP"),
        ("int",   "HP_PER_3LEVELS",       5,    "每3级+HP"),
    ]
    for r in rows4: ws4.append(r)
    auto_width(ws4)

    # --- #Spawn ---
    ws5 = wb.create_sheet("#Spawn")
    rows5 = [
        ("float", "RES_INTERVAL",       1.4,  "资源生成间隔"),
        ("int",   "RES_PER_SPAWN_MIN",  1,    "每次最少生成"),
        ("int",   "RES_PER_SPAWN_MAX",  2,    "每次最多生成"),
        ("int",   "RES_SIZE",           28,   "资源节点视觉大小"),
        ("float", "DECO_INTERVAL",      2.5,  "装饰物生成间隔"),
        ("int",   "DECO_PER_SPAWN",     1,    "每次装饰物数量"),
    ]
    for r in rows5: ws5.append(r)
    auto_width(ws5)

    # --- #Economy ---
    ws6 = wb.create_sheet("#Economy")
    rows6 = [
        ("int",   "BASE_TARGET",       13,    "基础提交目标"),
        ("float", "TARGET_GROWTH",     1.35,  "目标增长系数"),
        ("int",   "GOLD_PER_SUBMIT",   3,     "提交获得金币"),
        ("int",   "GOLD_PER_LEVEL",    25,    "升级获得金币"),
        ("int",   "LEVEL_DIST_TARGET", 1000,  "每关行驶目标距离(米)"),
        ("int",   "SUBMIT_BOX_W",      44,    "提交区域宽度"),
        ("int",   "SUBMIT_BOX_H",      44,    "提交区域高度"),
    ]
    for r in rows6: ws6.append(r)
    auto_width(ws6)

    wb.save(os.path.join(OUT_DIR, "GameConfig.xlsx"))
    print("  ✓ GameConfig.xlsx")


# =====================================================================
# 2. Turrets.xlsx — 炮塔数据表 + 投射物参数表
# =====================================================================
def make_turrets():
    wb = Workbook()

    # --- 数据表 CfgTurret ---
    ws = wb.active
    ws.title = "CfgTurret"
    # 三行表头
    ws.append(["编号","名称","图标Key","射程","伤害","冷却","颜色R","颜色G","颜色B","投射物类型"])
    ws.append(["int","string","string","int","int","float","int","int","int","string"])
    ws.append(["Id","Name","ImgKey","Range","Damage","Cooldown","ColorR","ColorG","ColorB","ProjType"])
    style_header(ws, 1, 10)
    style_header(ws, 2, 10)
    style_header(ws, 3, 10)
    data = [
        (1, "弓箭炮塔",   "arrow",    220, 8,  1.0,  180,140,80,  "arrow"),
        (2, "狙击炮塔",   "sniper",   350, 35, 5.0,  255,60,40,   "sniper"),
        (3, "喷火炮塔",   "flame",    200, 5,  0.12, 255,120,30,  "flame"),
        (4, "电能炮塔",   "electric", 200, 12, 0.7,  60,160,255,  "electric"),
        (5, "火箭炮塔",   "rocket",   280, 40, 8.0,  80,120,60,   "rocket"),
        (6, "机关枪炮塔", "minigun",  250, 3,  0.1,  255,230,80,  "minigun"),
    ]
    for d in data: ws.append(d)
    auto_width(ws)

    # --- 参数表 #Projectile ---
    ws2 = wb.create_sheet("#Projectile")
    proj_rows = [
        ("int",   "ARROW_SPEED",          280, "弓箭飞行速度"),
        ("float", "ARROW_LIFE",           1.5, "弓箭最大存活时间"),
        ("int",   "ARROW_HIT_RADIUS_SQ",  144, "弓箭命中判定距离²"),
        ("int",   "MINIGUN_SPEED",        500, "机关枪弹速"),
        ("float", "MINIGUN_LIFE",         0.8, "机关枪弹存活时间"),
        ("float", "MINIGUN_SPREAD",       0.12,"机关枪散布角度"),
        ("int",   "MINIGUN_HIT_RADIUS_SQ",144, "机关枪命中距离²"),
        ("int",   "ROCKET_SPEED",         150, "火箭飞行速度"),
        ("float", "ROCKET_LIFE",          3.0, "火箭最大存活时间"),
        ("int",   "ROCKET_AOE_RADIUS",    80,  "火箭爆炸半径(像素)"),
        ("float", "ROCKET_AOE_FALLOFF",   0.5, "火箭AOE边缘衰减比"),
        ("int",   "ROCKET_HIT_RADIUS_SQ", 225, "火箭命中判定距离²"),
        ("float", "ROCKET_TRAIL_INTERVAL",0.03,"火箭烟尾生成间隔"),
        ("int",   "SNIPER_SEGMENTS_PER10",1,   "狙击射线每10px段数"),
        ("float", "SNIPER_LIFE",          0.35,"狙击射线存活时间"),
        ("int",   "SNIPER_SPARK_COUNT",   6,   "狙击命中火花数"),
        ("float", "ELECTRIC_LIFE",        0.3, "闪电弧存活时间"),
        ("int",   "ELECTRIC_OFFSET",      22,  "闪电弧偏移量"),
        ("int",   "ELECTRIC_SPARK_COUNT", 4,   "闪电命中火花数"),
        ("float", "FLAME_HALF_SPREAD",    0.45,"火焰半扩散角(弧度)"),
        ("int",   "FLAME_DRAW_H",         130, "火焰视觉长度"),
    ]
    for r in proj_rows: ws2.append(r)
    auto_width(ws2)

    # --- 参数表 #TurretSlots ---
    ws3 = wb.create_sheet("#TurretSlots")
    slot_rows = [
        ("int", "TURRET_H",          48,  "炮塔渲染高度"),
        ("int", "TURRET_W",          36,  "炮塔渲染宽度"),
        ("int", "SLOT1_OFFSET_X",   -18,  "左侧槽位X偏移"),
        ("int", "SLOT1_OFFSET_Y",    6,   "左侧槽位Y偏移"),
        ("int", "SLOT2_OFFSET_X",    18,  "右侧槽位X偏移"),
        ("int", "SLOT2_OFFSET_Y",    6,   "右侧槽位Y偏移"),
        ("int", "SLOT3_OFFSET_X",    0,   "车头槽位X偏移"),
        ("int", "SLOT3_OFFSET_Y",    36,  "车头槽位Y偏移"),
        ("int", "SLOT4_OFFSET_X",    0,   "车尾槽位X偏移"),
        ("int", "SLOT4_OFFSET_Y",   -32,  "车尾槽位Y偏移"),
    ]
    for r in slot_rows: ws3.append(r)
    auto_width(ws3)

    wb.save(os.path.join(OUT_DIR, "Turrets.xlsx"))
    print("  ✓ Turrets.xlsx")


# =====================================================================
# 3. Resources.xlsx — 资源节点数据表
# =====================================================================
def make_resources():
    wb = Workbook()
    ws = wb.active
    ws.title = "CfgResource"
    ws.append(["编号","Key","名称","出现频率","生命值","掉落数量"])
    ws.append(["int","string","string","float","int","int"])
    ws.append(["Id","Key","Name","Freq","Hp","Drop"])
    style_header(ws, 1, 6)
    style_header(ws, 2, 6)
    style_header(ws, 3, 6)
    data = [
        (1, "wood",   "木材", 0.35, 80,  2),
        (2, "stone",  "岩石", 0.30, 100, 2),
        (3, "ore",    "矿石", 0.10, 150, 1),
        (4, "bush",   "灌木", 0.15, 50,  1),
        (5, "pebble", "碎石", 0.10, 60,  1),
    ]
    for d in data: ws.append(d)
    auto_width(ws)
    wb.save(os.path.join(OUT_DIR, "Resources.xlsx"))
    print("  ✓ Resources.xlsx")


# =====================================================================
# 4. Upgrades.xlsx — 肉鸽升级卡数据表
# =====================================================================
def make_upgrades():
    wb = Workbook()
    ws = wb.active
    ws.title = "CfgUpgrade"
    ws.append(["编号","Key","名称","描述","图标","是否炮塔卡","炮塔类型"])
    ws.append(["int","string","string","string","string","bool","string"])
    ws.append(["Id","Key","Name","Desc","Icon","IsTurret","TurretType"])
    style_header(ws, 1, 7)
    style_header(ws, 2, 7)
    style_header(ws, 3, 7)
    data = [
        (1,  "speed",    "急行军",   "移动速度+25%",       "boot",   False, ""),
        (2,  "carry",    "军用背包", "携带上限+5",         "bag",    False, ""),
        (3,  "atk",      "利刃强化", "攻击力+1",           "sword",  False, ""),
        (4,  "atkspd",   "疾风斩",   "攻击速度+30%",       "wind",   False, ""),
        (5,  "range",    "长臂猿",   "攻击范围+35%",       "magnet", False, ""),
        (6,  "gold",     "物资交换", "提交金币+50%",       "coin",   False, ""),
        (7,  "spawn",    "丰收区域", "资源生成+30%",       "star",   False, ""),
        (8,  "ore_luck", "探矿直觉", "矿石出现率翻倍",    "gem",    False, ""),
        (9,  "slow",     "制动系统", "列车速度-15%",       "gear",   False, ""),
        (10, "repair",   "应急维修", "列车回复30HP",       "shield", False, ""),
        (11, "armor",    "装甲强化", "列车最大HP+25",      "x2",     False, ""),
        (12, "multi",    "双倍搜刮", "采集掉落概率×2",     "fairy",  False, ""),
        (13, "turret_arrow",    "弓箭炮塔",   "解锁弓箭炮塔",     "turret_arrow",    True, "arrow"),
        (14, "turret_minigun",  "机关枪塔",   "解锁机关枪炮塔",   "turret_minigun",  True, "minigun"),
        (15, "turret_flame",    "喷火炮塔",   "解锁喷火炮塔",     "turret_flame",    True, "flame"),
        (16, "turret_sniper",   "狙击炮塔",   "解锁狙击炮塔",     "turret_sniper",   True, "sniper"),
        (17, "turret_electric", "电能炮塔",   "解锁电能炮塔",     "turret_electric", True, "electric"),
        (18, "turret_rocket",   "火箭炮塔",   "解锁火箭炮塔",     "turret_rocket",   True, "rocket"),
    ]
    for d in data: ws.append(d)
    auto_width(ws)
    wb.save(os.path.join(OUT_DIR, "Upgrades.xlsx"))
    print("  ✓ Upgrades.xlsx")


# =====================================================================
# 5. MetaData.xlsx — 局外系统(天赋/关卡/炮塔升级)
# =====================================================================
def make_metadata():
    wb = Workbook()

    # --- CfgTalent ---
    ws = wb.active
    ws.title = "CfgTalent"
    ws.append(["编号","Key","名称","描述","最大等级","基础费用","费用增长","图标"])
    ws.append(["int","string","string","string","int","int","float","string"])
    ws.append(["Id","Key","Name","Desc","MaxLv","CostBase","CostGrow","Icon"])
    style_header(ws, 1, 8)
    style_header(ws, 2, 8)
    style_header(ws, 3, 8)
    data = [
        (1, "hp",     "生命强化", "列车最大HP+10",     10, 50,  1.3, "image/talent_hp_20260421065056.png"),
        (2, "atk",    "力量提升", "攻击力+1",          10, 60,  1.3, "image/talent_atk_20260421065251.png"),
        (3, "atkspd", "疾速打击", "攻击速度+5%",       8,  80,  1.4, "image/talent_atkspd_20260421065058.png"),
        (4, "def",    "铁壁防御", "减少列车受伤-5%",   8,  70,  1.3, "image/talent_def_20260421065053.png"),
        (5, "speed",  "轻身术",   "移动速度+8%",       6,  60,  1.4, "image/talent_speed_20260421065223.png"),
        (6, "gold",   "聚财术",   "金币收益+10%",      8,  100, 1.5, "image/talent_gold_20260421065059.png"),
        (7, "carry",  "负重训练", "携带上限+2",        5,  80,  1.4, "image/talent_carry_20260421065057.png"),
        (8, "unlock", "求生本能", "解锁新技能",        3,  200, 2.0, "image/talent_unlock_20260421065109.png"),
    ]
    for d in data: ws.append(d)
    auto_width(ws)

    # --- CfgLevel ---
    ws2 = wb.create_sheet("CfgLevel")
    ws2.append(["编号","名称","波数","奖励金币","宝箱品质","是否解锁"])
    ws2.append(["int","string","int","int","string","bool"])
    ws2.append(["Id","Name","Waves","RewardGold","Chest","Unlocked"])
    style_header(ws2, 1, 6)
    style_header(ws2, 2, 6)
    style_header(ws2, 3, 6)
    data2 = [
        (1, "荒原前哨", 5,  50,  "bronze", True),
        (2, "冰封隧道", 7,  80,  "bronze", False),
        (3, "暴风雪谷", 8,  100, "silver", False),
        (4, "死寂车站", 10, 150, "silver", False),
        (5, "钢铁废墟", 12, 200, "gold",   False),
        (6, "末日核心", 15, 300, "gold",   False),
    ]
    for d in data2: ws2.append(d)
    auto_width(ws2)

    # --- CfgTurretUpgrade ---
    ws3 = wb.create_sheet("CfgTurretUpgrade")
    ws3.append(["编号","Key","名称","最大等级","碎片基数","碎片增长","图标"])
    ws3.append(["int","string","string","int","int","float","string"])
    ws3.append(["Id","Key","Name","MaxLv","FragBase","FragGrow","Icon"])
    style_header(ws3, 1, 7)
    style_header(ws3, 2, 7)
    style_header(ws3, 3, 7)
    data3 = [
        (1, "arrow",    "弓箭炮塔", 5, 5,  1.5, "image/turret_arrow_v3_20260420035036.png"),
        (2, "minigun",  "机关枪塔", 5, 8,  1.5, "image/turret_minigun_v3_20260420035022.png"),
        (3, "flame",    "喷火炮塔", 5, 8,  1.5, "image/edited_turret_flame_nofire_20260423065123.png"),
        (4, "sniper",   "狙击炮塔", 5, 10, 1.6, "image/turret_sniper_v3_20260420035021.png"),
        (5, "electric", "电能炮塔", 5, 10, 1.6, "image/turret_electric_v10_20260423040517.png"),
        (6, "rocket",   "火箭炮塔", 5, 12, 1.8, "image/turret_rocket_v3_20260420035019.png"),
    ]
    for d in data3: ws3.append(d)
    auto_width(ws3)

    wb.save(os.path.join(OUT_DIR, "MetaData.xlsx"))
    print("  ✓ MetaData.xlsx")


# =====================================================================
# 6. Equips.xlsx — 装备系统(装备槽位+品质+装备库)
# =====================================================================
def make_equips():
    wb = Workbook()

    # --- CfgEquipSlot ---
    ws = wb.active
    ws.title = "CfgEquipSlot"
    ws.append(["编号","Key","名称","图标"])
    ws.append(["int","string","string","string"])
    ws.append(["Id","Key","Name","Icon"])
    style_header(ws, 1, 4)
    style_header(ws, 2, 4)
    style_header(ws, 3, 4)
    data = [
        (1, "weapon",    "武器", "image/equip_weapon_sword_20260421064504.png"),
        (2, "accessory", "饰品", "image/equip_accessory_necklace_20260421064510.png"),
        (3, "ring",      "戒指", "image/equip_ring_20260421064511.png"),
        (4, "hat",       "帽子", "image/equip_hat_20260421064515.png"),
        (5, "clothes",   "衣服", "image/equip_clothes_20260421064729.png"),
        (6, "boots",     "鞋子", "image/equip_boots_20260421064742.png"),
    ]
    for d in data: ws.append(d)
    auto_width(ws)

    # --- CfgQuality ---
    ws2 = wb.create_sheet("CfgQuality")
    ws2.append(["编号","Key","名称","颜色R","颜色G","颜色B","颜色A"])
    ws2.append(["int","string","string","int","int","int","int"])
    ws2.append(["Id","Key","Name","ColorR","ColorG","ColorB","ColorA"])
    style_header(ws2, 1, 7)
    style_header(ws2, 2, 7)
    style_header(ws2, 3, 7)
    data2 = [
        (1, "common",   "普通", 160, 165, 175, 255),
        (2, "uncommon", "优秀", 65,  170, 80,  255),
        (3, "rare",     "稀有", 55,  120, 210, 255),
        (4, "epic",     "史诗", 150, 60,  200, 255),
        (5, "legend",   "传说", 220, 165, 30,  255),
    ]
    for d in data2: ws2.append(d)
    auto_width(ws2)

    # --- CfgEquip ---
    ws3 = wb.create_sheet("CfgEquip")
    ws3.append(["编号","Key","槽位","名称","品质","攻击力","防御力","攻速","生命","速度","图标"])
    ws3.append(["int","string","string","string","int","int","int","float","int","float","string"])
    ws3.append(["Id","Key","Slot","Name","Quality","Atk","Def","AtkSpd","Hp","Speed","Icon"])
    style_header(ws3, 1, 11)
    style_header(ws3, 2, 11)
    style_header(ws3, 3, 11)
    data3 = [
        (1, "sword_iron",  "weapon",    "铁剑",   1, 5,0,0,  0,0,   "image/equip_weapon_sword_20260421064504.png"),
        (2, "axe_war",     "weapon",    "战斧",   2, 8,0,0,  0,0,   "image/equip_weapon_axe_20260421064553.png"),
        (3, "necklace_1",  "accessory", "护身符", 1, 0,3,0,  0,0,   "image/equip_accessory_necklace_20260421064510.png"),
        (4, "ring_silver", "ring",      "银戒指", 2, 0,0,0.1,0,0,   "image/equip_ring_20260421064511.png"),
        (5, "hat_winter",  "hat",       "冬帽",   1, 0,0,0,  10,0,  "image/equip_hat_20260421064515.png"),
        (6, "coat_warm",   "clothes",   "棉衣",   1, 0,5,0,  0,0,   "image/equip_clothes_20260421064729.png"),
        (7, "boots_army",  "boots",     "军靴",   1, 0,0,0,  0,0.1, "image/equip_boots_20260421064742.png"),
    ]
    for d in data3: ws3.append(d)
    auto_width(ws3)

    wb.save(os.path.join(OUT_DIR, "Equips.xlsx"))
    print("  ✓ Equips.xlsx")


# =====================================================================
# 7. Shop.xlsx — 商城系统
# =====================================================================
def make_shop():
    wb = Workbook()

    # --- CfgShopDaily ---
    ws = wb.active
    ws.title = "CfgShopDaily"
    ws.append(["编号","Key","名称","描述","价格","货币类型","图标"])
    ws.append(["int","string","string","string","int","string","string"])
    ws.append(["Id","Key","Name","Desc","Price","Currency","Icon"])
    style_header(ws, 1, 7)
    style_header(ws, 2, 7)
    style_header(ws, 3, 7)
    data = [
        (1, "daily_gold",  "金币礼包", "获得500金币",      50,  "diamond", "image/hud_gold_coin.png"),
        (2, "daily_wood",  "木材补给", "获得20木材",       30,  "diamond", "image/meta_icon_wood_20260421063709.png"),
        (3, "daily_stone", "石材补给", "获得15石材",       30,  "diamond", "image/meta_icon_stone_20260421063705.png"),
        (4, "daily_chest", "银色宝箱", "随机稀有装备×1",   100, "diamond", "image/chest_silver_20260421064733.png"),
    ]
    for d in data: ws.append(d)
    auto_width(ws)

    # --- #Gacha ---
    ws2 = wb.create_sheet("#Gacha")
    rows2 = [
        ("string", "NAME",        "末日抽卡",               "抽卡系统名称"),
        ("string", "DESC",        "消耗钻石抽取装备碎片",   "抽卡描述"),
        ("int",    "COST_SINGLE", 100,                      "单抽费用"),
        ("int",    "COST_TEN",    900,                      "十连费用"),
    ]
    for r in rows2: ws2.append(r)
    auto_width(ws2)

    wb.save(os.path.join(OUT_DIR, "Shop.xlsx"))
    print("  ✓ Shop.xlsx")


# =====================================================================
# 8. MetaUI.xlsx — 局外UI/Tab/货币/宝箱/初始存档
# =====================================================================
def make_meta_ui():
    wb = Workbook()

    # --- CfgTab ---
    ws = wb.active
    ws.title = "CfgTab"
    ws.append(["编号","Key","名称","图标"])
    ws.append(["int","string","string","string"])
    ws.append(["Id","Key","Name","Icon"])
    style_header(ws, 1, 4)
    style_header(ws, 2, 4)
    style_header(ws, 3, 4)
    data = [
        (1, "shop",   "商城", "image/meta_tab_shop_20260421063707.png"),
        (2, "equip",  "装备", "image/meta_tab_equip_20260421063710.png"),
        (3, "battle", "战斗", "image/meta_tab_battle_20260421063808.png"),
        (4, "train",  "列车", "image/meta_tab_train_20260421063842.png"),
        (5, "talent", "天赋", "image/meta_tab_talent_20260421063732.png"),
    ]
    for d in data: ws.append(d)
    auto_width(ws)

    # --- CfgCurrency ---
    ws2 = wb.create_sheet("CfgCurrency")
    ws2.append(["编号","Key","图标"])
    ws2.append(["int","string","string"])
    ws2.append(["Id","Key","Icon"])
    style_header(ws2, 1, 3)
    style_header(ws2, 2, 3)
    style_header(ws2, 3, 3)
    data2 = [
        (1, "gold",    "image/hud_gold_coin.png"),
        (2, "diamond", "image/meta_icon_diamond_20260421063718.png"),
        (3, "wood",    "image/meta_icon_wood_20260421063709.png"),
        (4, "stone",   "image/meta_icon_stone_20260421063705.png"),
    ]
    for d in data2: ws2.append(d)
    auto_width(ws2)

    # --- CfgChest ---
    ws3 = wb.create_sheet("CfgChest")
    ws3.append(["编号","Key","图标"])
    ws3.append(["int","string","string"])
    ws3.append(["Id","Key","Icon"])
    style_header(ws3, 1, 3)
    style_header(ws3, 2, 3)
    style_header(ws3, 3, 3)
    data3 = [
        (1, "bronze", "image/chest_bronze_20260421064750.png"),
        (2, "silver", "image/chest_silver_20260421064733.png"),
        (3, "gold",   "image/chest_gold_20260421064736.png"),
    ]
    for d in data3: ws3.append(d)
    auto_width(ws3)

    # --- #InitSave ---
    ws4 = wb.create_sheet("#InitSave")
    rows4 = [
        ("int",    "INIT_GOLD",      500,    "初始金币"),
        ("int",    "INIT_DIAMOND",   100,    "初始钻石"),
        ("int",    "INIT_WOOD",      30,     "初始木材"),
        ("int",    "INIT_STONE",     20,     "初始石材"),
        ("int",    "INIT_MAX_LEVEL", 1,      "初始最高解锁关卡"),
        ("string", "INIT_PLAYER_NAME","幸存者","初始玩家名"),
        ("int",    "INIT_PLAYER_LEVEL",1,     "初始玩家等级"),
    ]
    for r in rows4: ws4.append(r)
    auto_width(ws4)

    wb.save(os.path.join(OUT_DIR, "MetaUI.xlsx"))
    print("  ✓ MetaUI.xlsx")


# =====================================================================
# 9. Colors.xlsx — 全部颜色主题(游戏内 + 局外UI)
# =====================================================================
def make_colors():
    wb = Workbook()

    # --- #GameColors (游戏内颜色) ---
    ws = wb.active
    ws.title = "#GameColors"
    game_colors = [
        # 雪地
        ("luatable", "SNOW1",         "{ 155, 160, 168 }",   "脏灰雪"),
        ("luatable", "SNOW2",         "{ 130, 138, 148 }",   "深灰雪纹"),
        ("luatable", "SNOW3",         "{ 170, 175, 182 }",   "浅灰"),
        ("luatable", "SNOW_DARK",     "{ 105, 112, 125 }",   "阴影区"),
        # 道路
        ("luatable", "PATH",          "{ 85, 78, 68 }",      "暗褐冻土"),
        ("luatable", "PATH_LIGHT",    "{ 100, 92, 80 }",     "碎石亮面"),
        ("luatable", "PATH_DARK",     "{ 60, 55, 45 }",      "泥坑暗部"),
        ("luatable", "PATH_EDGE",     "{ 120, 125, 135 }",   "路缘过渡"),
        # 列车
        ("luatable", "TRAIN_BODY",    "{ 52, 55, 50 }",      "暗铁灰车身"),
        ("luatable", "TRAIN_DARK",    "{ 35, 38, 32 }",      "深暗面"),
        ("luatable", "TRAIN_LIGHT",   "{ 72, 78, 68 }",      "微弱高光"),
        ("luatable", "TRAIN_ROOF",    "{ 40, 42, 38 }",      "车顶"),
        ("luatable", "TRAIN_METAL",   "{ 65, 68, 62 }",      "金属件"),
        ("luatable", "TRAIN_WINDOW",  "{ 75, 105, 130 }",    "脏窗玻璃"),
        ("luatable", "TRAIN_WINDOW2", "{ 55, 80, 105 }",     "深色窗"),
        ("luatable", "TRAIN_STRIPE",  "{ 140, 35, 25 }",     "暗血红装饰"),
        ("luatable", "TRAIN_STRIPE2", "{ 110, 25, 18 }",     "深红锈"),
        ("luatable", "TRAIN_BRASS",   "{ 145, 120, 55 }",    "氧化铜"),
        ("luatable", "TRAIN_BRASS2",  "{ 170, 145, 70 }",    "铜亮面"),
        ("luatable", "TRAIN_BRASS3",  "{ 110, 90, 40 }",     "铜暗面"),
        ("luatable", "TRAIN_SMOKE",   "{ 38, 35, 30 }",      "烟囱黑"),
        ("luatable", "TRAIN_SMOKE2",  "{ 25, 22, 18 }",      "烟囱暗"),
        ("luatable", "TRAIN_RIVET",   "{ 80, 85, 75 }",      "锈铆钉"),
        ("luatable", "TRAIN_FRAME",   "{ 30, 32, 28 }",      "暗轮廓"),
        ("luatable", "TRAIN_PLATE",   "{ 120, 105, 55 }",    "锈铭牌"),
        ("luatable", "TRAIN_CAB_BG",  "{ 45, 48, 42 }",      "驾驶室"),
        ("luatable", "TRAIN_PIPE",    "{ 95, 90, 80 }",      "锈管道"),
        # 铁轨
        ("luatable", "RAIL_COLOR",    "{ 75, 78, 85 }",      "暗铁轨"),
        ("luatable", "RAIL_LIGHT",    "{ 100, 105, 115 }",   "磨亮面"),
        ("luatable", "SLEEPER_COLOR", "{ 70, 60, 42 }",      "腐朽枕木"),
        ("luatable", "SLEEPER_DARK",  "{ 48, 40, 28 }",      "枕木暗面"),
        ("luatable", "BALLAST_COLOR", "{ 80, 75, 65 }",      "暗碎石"),
        # 提交方块
        ("luatable", "SUBMIT_BG",     "{ 55, 58, 65, 220 }", "提交区背景"),
        ("luatable", "SUBMIT_BORDER", "{ 85, 90, 100 }",     "提交区边框"),
        ("luatable", "SUBMIT_FULL",   "{ 50, 130, 80 }",     "已满暗绿"),
        # HUD
        ("luatable", "HUD_BG",        "{ 18, 20, 28, 240 }", "HUD背景"),
        ("luatable", "HUD_BORDER",    "{ 40, 45, 58 }",      "HUD边框"),
        ("luatable", "HUD_TEXT",       "{ 175, 180, 190 }",   "HUD文字"),
        # 树木
        ("luatable", "TRUNK",         "{ 55, 45, 35 }",      "黑褐树干"),
        ("luatable", "TRUNK_DARK",    "{ 35, 28, 20 }",      "腐烂暗部"),
        ("luatable", "TREE_SNOW",     "{ 140, 148, 160 }",   "灰雪"),
        ("luatable", "TREE_BARE",     "{ 65, 52, 40 }",      "枯枝"),
        ("luatable", "TREE_DARK",     "{ 45, 38, 28 }",      "暗枝"),
        # 雪堆/残骸
        ("luatable", "SNOWDRIFT1",    "{ 148, 155, 165 }",   "脏雪堆"),
        ("luatable", "SNOWDRIFT2",    "{ 120, 128, 140 }",   "深灰雪"),
        ("luatable", "RUIN_COLOR",    "{ 68, 62, 55 }",      "废墟残骸"),
        # 资源
        ("luatable", "WOOD_COLOR",    "{ 110, 78, 40 }",     "暗木色"),
        ("luatable", "STONE_COLOR",   "{ 95, 95, 105 }",     "暗岩灰"),
        ("luatable", "ORE_COLOR",     "{ 65, 120, 170 }",    "深蓝矿脉"),
        ("luatable", "BUSH_COLOR",    "{ 60, 100, 50 }",     "深绿灌木"),
        ("luatable", "PEBBLE_COLOR",  "{ 130, 125, 115 }",   "浅灰碎石"),
        ("luatable", "GOLD_COLOR",    "{ 200, 165, 40 }",    "暗金"),
        # 丧尸
        ("luatable", "ZOMBIE_SKIN",   "{ 85, 110, 72 }",     "腐肉绿"),
        ("luatable", "ZOMBIE_DARK",   "{ 50, 65, 38 }",      "深腐色"),
        ("luatable", "ZOMBIE_EYE",    "{ 200, 35, 20 }",     "血红眼"),
        ("luatable", "ZOMBIE_CLOTH",  "{ 55, 48, 42 }",      "破烂深色"),
        # 玩家
        ("luatable", "PLAYER_COAT",   "{ 125, 55, 35 }",     "暗红破旧大衣"),
        ("luatable", "PLAYER_SKIN",   "{ 185, 155, 130 }",   "偏暗肤色"),
        ("luatable", "PLAYER_PANTS",  "{ 48, 50, 58 }",      "深灰裤"),
        ("luatable", "PLAYER_BOOT",   "{ 38, 35, 28 }",      "泥靴"),
        ("luatable", "PLAYER_HAT",    "{ 105, 42, 28 }",     "暗红帽"),
        ("luatable", "PLAYER_SCARF",  "{ 145, 125, 40 }",    "脏黄围巾"),
        # 文字
        ("luatable", "TEXT_WHITE",    "{ 220, 225, 230 }",   "偏灰白"),
        ("luatable", "TEXT_DARK",     "{ 15, 18, 25 }",      "深色"),
        ("luatable", "TEXT_DIM",      "{ 100, 108, 120 }",   "暗灰"),
        ("luatable", "TEXT_RED",      "{ 190, 50, 35 }",     "暗血红"),
        # 卡片
        ("luatable", "CARD_BG",       "{ 28, 32, 42 }",      "卡片背景"),
        ("luatable", "CARD_BORDER",   "{ 55, 60, 75 }",      "卡片边框"),
        ("luatable", "CARD_GLOW",     "{ 70, 140, 200 }",    "卡片发光"),
        # 飙血
        ("luatable", "BLOOD_COLOR",   "{ 180, 30, 20 }",     "飙血颜色"),
    ]
    for r in game_colors: ws.append(r)
    auto_width(ws)

    # --- #MetaUIColors (局外UI颜色) ---
    ws2 = wb.create_sheet("#MetaUIColors")
    meta_colors = [
        # 主背景
        ("luatable", "BG_DARK",        "{ 18, 20, 28, 255 }",     "最深背景"),
        ("luatable", "BG_PANEL",       "{ 28, 32, 42, 255 }",     "面板背景"),
        ("luatable", "BG_CARD",        "{ 38, 42, 55, 255 }",     "卡片背景"),
        ("luatable", "BG_INPUT",       "{ 22, 25, 35, 255 }",     "输入框背景"),
        # 顶栏/底栏
        ("luatable", "BAR_BG",         "{ 14, 16, 22, 245 }",     "顶/底栏背景"),
        ("luatable", "BAR_BORDER",     "{ 45, 50, 65, 255 }",     "栏边框"),
        # Tab
        ("luatable", "TAB_ACTIVE",     "{ 70, 140, 200, 255 }",   "激活Tab颜色"),
        ("luatable", "TAB_INACTIVE",   "{ 100, 108, 120, 255 }",  "未激活Tab"),
        ("luatable", "TAB_BG",         "{ 22, 25, 35, 240 }",     "Tab背景"),
        # 按钮
        ("luatable", "BTN_PRIMARY",    "{ 55, 130, 85, 255 }",    "主按钮暗绿"),
        ("luatable", "BTN_SECONDARY",  "{ 50, 55, 70, 255 }",     "次按钮"),
        ("luatable", "BTN_DANGER",     "{ 160, 45, 35, 255 }",    "危险按钮"),
        ("luatable", "BTN_GOLD",       "{ 180, 145, 40, 255 }",   "金色按钮"),
        # 文字
        ("luatable", "META_TEXT_WHITE","{ 220, 225, 230, 255 }",  "白色文字"),
        ("luatable", "META_TEXT_GRAY", "{ 130, 138, 150, 255 }",  "灰色文字"),
        ("luatable", "META_TEXT_DIM",  "{ 80, 85, 95, 255 }",     "暗灰文字"),
        ("luatable", "META_TEXT_GOLD", "{ 220, 185, 50, 255 }",   "金色文字"),
        ("luatable", "META_TEXT_GREEN","{ 65, 180, 90, 255 }",    "绿色文字"),
        ("luatable", "META_TEXT_RED",  "{ 200, 55, 40, 255 }",    "红色文字"),
        # 分割线
        ("luatable", "DIVIDER",        "{ 40, 45, 58, 255 }",     "分割线"),
    ]
    for r in meta_colors: ws2.append(r)
    auto_width(ws2)

    wb.save(os.path.join(OUT_DIR, "Colors.xlsx"))
    print("  ✓ Colors.xlsx")


# =====================================================================
# Main
# =====================================================================
if __name__ == "__main__":
    print("=== 生成完整 Excel 配置表 ===")
    # 清空旧文件
    for f in os.listdir(OUT_DIR):
        if f.endswith(".xlsx"):
            os.remove(os.path.join(OUT_DIR, f))
            print(f"  删除旧文件: {f}")

    make_game_config()
    make_turrets()
    make_resources()
    make_upgrades()
    make_metadata()
    make_equips()
    make_shop()
    make_meta_ui()
    make_colors()
    print(f"\n=== 完成！共 9 个 Excel 文件在 {OUT_DIR} ===")
